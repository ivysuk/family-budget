"""가족 공동 통장 관리용 구글시트 초기 구조(탭+헤더)를 만든다.
로컬에서 xlsx로 만든 뒤 구글드라이브에 업로드해서 네이티브 구글시트로 변환한다.
"""
from openpyxl import Workbook
from openpyxl.styles import Font

wb = Workbook()

sheets = {
    "가족구성원": ["이름", "연락처"],
    "월별납부": ["연월", "이름", "납부금액", "납부일자", "상태"],
    "지출청구": ["청구ID", "청구일자", "청구자", "항목", "금액", "메모", "정산상태", "이체일자"],
    "렌탈료(정기지출)": ["항목명", "정기금액", "주기", "다음청구예정일"],
    "거래원장": ["날짜", "차변계정", "대변계정", "금액", "관련인물", "메모", "출처유형", "출처ID"],
    "잔액대사": ["연월", "시스템잔액(자동계산)", "실제통장잔액(수동입력)", "차이"],
}

first = True
for name, headers in sheets.items():
    ws = wb.active if first else wb.create_sheet()
    ws.title = name
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 16
    first = False

wb.save("family_budget_template.xlsx")
print("saved family_budget_template.xlsx")
